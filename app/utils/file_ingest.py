from __future__ import annotations

# Security-log-analysis mainline file-ingest helpers.

import csv
import io
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from app.models.event import RawEvent


PROJECT_DOC_FILENAME_TOKENS = (
    "prd",
    "说明",
    "文档",
    "设计",
    "字段定义",
    "任务拆分",
    "合同文档",
    "原型",
    "需求",
)

PROJECT_DOC_CONTENT_TOKENS = (
    "接口",
    "字段",
    "页面",
    "原型",
    "数据库",
    "需求",
    "任务拆分",
    "产品说明",
    "开发计划",
)


def _safe_decode(data: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    text = "\n\n".join(part for part in pages if part)
    if not text:
        raise ValueError("PDF 已上传，但没有提取到可分析文本。")
    return text


def extract_xlsx_payload(data: bytes, filename: str) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX 已上传，但没有解析出可分析内容。")
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = []
    seen_headers: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_tabular_header(header) or f"column_{index + 1}"
        seen_headers[normalized] = seen_headers.get(normalized, 0) + 1
        if seen_headers[normalized] > 1:
            normalized = f"{normalized}_{seen_headers[normalized]}"
        normalized_headers.append(normalized)
    parsed_rows: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        if raw_row is None:
            continue
        row = {
            normalized_headers[index]: ("" if value is None else value)
            for index, value in enumerate(raw_row[: len(normalized_headers)])
        }
        if any(str(value).strip() for value in row.values()):
            parsed_rows.append(row)
    if not parsed_rows:
        raise ValueError("XLSX 已上传，但没有解析出有效数据行。")
    sample_lines = []
    for row in parsed_rows[:5]:
        sample_lines.append(" ".join(str(value) for value in row.values() if value not in {"", None}))
    return {
        "headers": normalized_headers,
        "rows": parsed_rows,
        "row_count": len(parsed_rows),
        "parser_profile": "xlsx-tabular",
        "filename": filename,
        "sheet_name": worksheet.title,
        "content_preview": "\n".join(sample_lines),
    }


def infer_source_type(filename: str, content: str, parser_profile: str | None = None) -> str:
    lower_name = filename.lower()
    lower_content = content.lower()
    tabular_like = parser_profile in {"xlsx-tabular", "csv-tabular"}
    if parser_profile in {"text", "pdf-text", "json"} and (
        any(token in lower_name for token in PROJECT_DOC_FILENAME_TOKENS)
        or sum(1 for token in PROJECT_DOC_CONTENT_TOKENS if token in lower_content) >= 2
    ):
        return "github"
    if any(token in lower_name for token in ("operatelog", "operationlog", "operate_log")) or (
        tabular_like and any(
        token in lower_content for token in ("组织名称", "资源类型", "动作", "创建用户会话", "导出")
        )
    ):
        return "operation_audit"
    if any(token in lower_name for token in ("jumpserver", "userloginlog", "loginlog")) or (
        tabular_like and any(
        token in lower_content for token in ("登录 ip", "登录日期", "认证方式", "原因描述", "otp")
        )
    ):
        return "login_auth"
    if any(token in lower_name for token in ("ftplog", "sftp", "upload", "download")) or (
        tabular_like and any(
        token in lower_content for token in ("文件名", "开始日期", "可下载", "操作")
        )
    ):
        return "file_transfer_audit"
    if any(token in lower_name for token in ("command", "cmdlog", "commandlog")) or (
        tabular_like and any(
        token in lower_content for token in ("会话", "远端地址", "风险等级", "命令")
        )
    ):
        return "command_audit"
    if any(token in lower_name for token in ("incident", "edr", "endpoint")) or any(token in lower_content for token in ("attack.localfileinclusion", "endpoint", "shellspawned", "credentialtheft")):
        return "endpoint"
    if any(token in lower_name for token in ("riskanalytics", "baseline", "hardening", "audit")) or any(token in lower_content for token in ("发现名称", "风险评分", "auditd", "aide", "ssh 访问未受限制")):
        return "host_risk_analytics"
    if any(token in lower_name for token in ("cloud", "iam", "bucket")) or any(token in lower_content for token in ("cloudtrail", "security group", "public bucket")):
        return "cloud"
    if any(token in lower_name for token in ("github", "pr", "commit")) or "pull_request" in lower_content:
        return "github"
    if any(token in lower_name for token in ("nmap", "asset", "tls")) or any(token in lower_content for token in ("port", "certificate", "subdomain")):
        return "easm"
    if "kms" in lower_name or "sign" in lower_content:
        return "kms"
    return "github"


def parse_file_to_raw_event(filename: str, data: bytes) -> RawEvent:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf" or data.startswith(b"%PDF-"):
        content = extract_pdf_text(data)
        payload = {"content": content, "parser_profile": "pdf-text", "filename": filename}
        source_type = infer_source_type(filename, content, "pdf-text")
    elif suffix == ".json":
        content = _safe_decode(data)
        try:
            payload = json.loads(content)
            if not isinstance(payload, dict):
                payload = {"content": content}
        except json.JSONDecodeError:
            payload = {"content": content}
        payload["parser_profile"] = "json"
        payload["filename"] = filename
        source_type = infer_source_type(filename, content, "json")
    elif suffix == ".csv":
        content = _safe_decode(data)
        reader = csv.DictReader(io.StringIO(content))
        rows = [dict(row) for row in reader]
        payload = {
            "headers": list(reader.fieldnames or []),
            "rows": rows[:50],
            "row_count": len(rows),
            "parser_profile": "csv-tabular",
            "filename": filename,
        }
        source_type = infer_source_type(filename, content, "csv-tabular")
    elif suffix in {".xlsx", ".xlsm"}:
        payload = extract_xlsx_payload(data, filename)
        preview = " ".join(payload.get("headers", [])) + "\n" + str(payload.get("content_preview") or "")
        source_type = infer_source_type(filename, preview, "xlsx-tabular")
    else:
        content = _safe_decode(data)
        payload = {"content": content, "parser_profile": "text", "filename": filename}
        source_type = infer_source_type(filename, content, "text")
    stem = Path(filename).stem
    return RawEvent(
        source_type=source_type,
        asset_context={
            "asset_id": stem,
            "asset_name": stem,
            "environment": "unknown",
            "criticality": 3,
            "source_file": filename,
            "source_type": source_type,
        },
        payload=payload,
    )


def _string_value(value: Any) -> str:
    return str(value or "").strip()


def _normalize_tabular_header(value: Any) -> str:
    text = _string_value(value)
    while text.startswith("*"):
        text = text[1:].strip()
    return text


def _lower_value(value: Any) -> str:
    return _string_value(value).lower()


def _parse_success_flag(value: Any) -> bool:
    lowered = _lower_value(value)
    return lowered in {"true", "yes", "1", "成功", "success", "已成功"}


def _count_keywords(command: str, patterns: dict[str, tuple[str, ...]]) -> dict[str, int]:
    lowered = command.lower()
    return {key: int(any(token in lowered for token in tokens)) for key, tokens in patterns.items()}


def build_jumpserver_composite_raw_event(raw_events: list[RawEvent]) -> RawEvent | None:
    login_events = [event for event in raw_events if event.source_type == "login_auth"]
    command_events = [event for event in raw_events if event.source_type == "command_audit"]
    transfer_events = [event for event in raw_events if event.source_type == "file_transfer_audit"]
    operation_events = [event for event in raw_events if event.source_type == "operation_audit"]
    if not command_events or not (login_events or transfer_events or operation_events):
        return None

    latest_login_event = None
    supplemental_login_events: list[RawEvent] = []
    if login_events:
        ordered_login_events = sorted(
            login_events,
            key=lambda event: str(event.asset_context.get("source_file") or ""),
        )
        latest_login_event = ordered_login_events[-1]
        supplemental_login_events = ordered_login_events[:-1]

    login_rows = [row for row in (latest_login_event.payload.get("rows", []) if latest_login_event else []) if isinstance(row, dict)]
    command_rows = [row for event in command_events for row in event.payload.get("rows", []) if isinstance(row, dict)]
    transfer_rows = [row for event in transfer_events for row in event.payload.get("rows", []) if isinstance(row, dict)]
    operation_rows = [row for event in operation_events for row in event.payload.get("rows", []) if isinstance(row, dict)]

    login_success = 0
    login_failures = 0
    failed_accounts: dict[str, int] = {}
    proxy_ips: set[str] = set()
    for row in login_rows:
        status = _lower_value(row.get("状态") or row.get("status"))
        user = _string_value(row.get("用户名") or row.get("user"))
        ip = _string_value(row.get("登录 IP") or row.get("src_ip"))
        if ip.startswith("10."):
            proxy_ips.add(ip)
        if any(token in status for token in ("成功", "success")):
            login_success += 1
        else:
            login_failures += 1
            if user:
                failed_accounts[user] = failed_accounts.get(user, 0) + 1

    command_patterns = {
        "privilege_escalation": ("sudo", "sudo su", "sudo su -"),
        "service_control": ("systemctl start", "systemctl stop", "systemctl restart", "systemctl status"),
        "download_and_execute": ("wget ", "curl ", "| bash", "install.sh"),
        "file_replace": ("rm -rf", "mv ", "ln -sf"),
        "permission_relaxation": ("chmod", "chown"),
        "lateral_movement": ("rsync", "scp ", "ssh ", "telnet ", "nc "),
        "binary_execution": ("./", "/tmp/"),
        "runner_driven_batch_change": ("hot-upgrade", "patch.sh", "runner"),
    }
    noisy_markers = ("\u001b", "tmux", "claude", "codex")
    command_pattern_labels = {
        "privilege_escalation": "提权",
        "service_control": "服务控制",
        "download_and_execute": "下载执行",
        "file_replace": "删除替换",
        "permission_relaxation": "权限变更",
        "lateral_movement": "横向/远程操作",
        "binary_execution": "本地执行",
        "runner_driven_batch_change": "批量发布/补丁分发",
    }
    account_stats: dict[str, dict[str, Any]] = {}
    effective_command_count = 0
    command_findings: dict[str, int] = {key: 0 for key in command_patterns}
    for row in command_rows:
        command = _string_value(row.get("命令") or row.get("command"))
        if not command or any(marker in command.lower() for marker in noisy_markers):
            continue
        effective_command_count += 1
        user = _string_value(row.get("用户") or row.get("user"))
        asset = _string_value(row.get("资产") or row.get("asset"))
        stats = account_stats.setdefault(
            user or "unknown",
            {
                "user": user or "unknown",
                "assets": set(),
                "risk_actions": set(),
                "risk_counts": {key: 0 for key in command_patterns},
                "representative_commands": [],
                "supporting_sources": {"command_audit"},
                "effective_command_count": 0,
            },
        )
        if asset:
            stats["assets"].add(asset)
        stats["effective_command_count"] += 1
        if len(stats["representative_commands"]) < 8 and command not in stats["representative_commands"]:
            stats["representative_commands"].append(command)
        matches = _count_keywords(command, command_patterns)
        for risk_action, hit in matches.items():
            if hit:
                command_findings[risk_action] += 1
                stats["risk_actions"].add(risk_action)
                stats["risk_counts"][risk_action] += 1

    transfer_uploads = []
    transfer_downloads = []
    transfer_findings = {
        "binary_drop_to_tmp": 0,
        "data_export": 0,
    }
    for row in transfer_rows:
        user = _string_value(row.get("用户") or row.get("user"))
        asset = _string_value(row.get("资产") or row.get("asset"))
        filename = _string_value(row.get("文件名") or row.get("filename"))
        operation = _lower_value(row.get("操作") or row.get("operation"))
        account = _string_value(row.get("账号") or row.get("account"))
        record = {
            "user": user,
            "asset": asset,
            "filename": filename,
            "operation": operation,
            "account": account,
            "event_time": _string_value(row.get("开始日期") or row.get("event_time")),
        }
        if any(token in operation for token in ("上传", "upload")):
            transfer_uploads.append(record)
            if "/tmp/" in filename.lower() or filename.lower().startswith("tmp/") or account.lower() == "root":
                transfer_findings["binary_drop_to_tmp"] += 1
        elif any(token in operation for token in ("下载", "download")):
            transfer_downloads.append(record)
            transfer_findings["data_export"] += 1

    operation_actions: dict[str, int] = {}
    operator_counts: dict[str, int] = {}
    resource_type_counts: dict[str, int] = {}
    operation_findings = {
        "export_actions": 0,
        "host_or_account_creation": 0,
        "authorization_updates": 0,
        "session_creation": 0,
    }
    high_impact_operations = []
    export_events = []
    for row in operation_rows:
        user = _string_value(row.get("用户") or row.get("user"))
        action = _string_value(row.get("动作") or row.get("action"))
        resource_type = _string_value(row.get("资源类型") or row.get("resource_type"))
        resource = _string_value(row.get("资源") or row.get("resource"))
        event_time = _string_value(row.get("日期") or row.get("event_time"))
        lowered = action.lower()
        operation_actions[action or "unknown"] = operation_actions.get(action or "unknown", 0) + 1
        if user:
            operator_counts[user] = operator_counts.get(user, 0) + 1
        if resource_type:
            resource_type_counts[resource_type] = resource_type_counts.get(resource_type, 0) + 1
        if any(token in lowered for token in ("导出", "export")):
            operation_findings["export_actions"] += 1
            export_events.append(
                {
                    "event_time": event_time,
                    "user": user,
                    "resource_type": resource_type,
                    "resource": resource,
                }
            )
        if any(token in lowered for token in ("创建用户会话", "session")):
            operation_findings["session_creation"] += 1
        if any(token in lowered for token in ("创建", "create")) and any(token in resource_type.lower() for token in ("主机", "资产", "账号", "用户")):
            operation_findings["host_or_account_creation"] += 1
        if any(token in lowered for token in ("更新", "授权", "update", "grant", "bind")):
            operation_findings["authorization_updates"] += 1
        if any(
            count for count in [
                any(token in lowered for token in ("导出", "export")),
                any(token in lowered for token in ("创建", "create", "更新", "授权", "grant", "bind")),
            ]
        ):
            high_impact_operations.append(
                {
                    "user": user,
                    "action": action,
                    "resource_type": resource_type,
                    "resource": resource,
                    "event_time": event_time,
                }
            )

    cross_source_correlations = []
    for upload in transfer_uploads:
        user = upload["user"]
        asset = upload["asset"]
        related = account_stats.get(user or "unknown")
        if not related:
            continue
        command_blob = " ".join(related["representative_commands"]).lower()
        chain = ["file_upload"]
        if "sudo" in command_blob:
            chain.append("privilege_escalation")
        if any(token in command_blob for token in ("chmod", "chown")):
            chain.append("chmod_or_move")
        if any(token in command_blob for token in ("./", "mv /tmp", "mv ", "/tmp/")):
            chain.append("binary_execution")
        if any(token in command_blob for token in ("systemctl", "curl ", "telnet", "nc ", "rsync", "scp ", "ssh ")):
            chain.append("service_change_or_network_test")
        if len(chain) >= 3:
            cross_source_correlations.append(
                {
                    "user": user,
                    "asset": asset,
                    "time_window": "same upload batch",
                    "correlation_chain": chain,
                    "confidence": "high" if len(chain) >= 5 else "medium",
                    "judgement": "形成了从文件投放到执行与验证的高风险操作链，需要人工复核。",
                }
            )
    if high_impact_operations and cross_source_correlations:
        for item in cross_source_correlations[:]:
            item["control_plane_support"] = "存在导出/创建/授权/会话创建等管理平面证据，可用于解释主机侧链条背景，但不能替代主机侧执行证据。"

    ranked_accounts = sorted(
        account_stats.values(),
        key=lambda item: (len(item["risk_actions"]), item["effective_command_count"], len(item["assets"])),
        reverse=True,
    )
    high_risk_accounts = []
    for item in ranked_accounts[:8]:
        user = item["user"]
        supporting_sources = set(item["supporting_sources"])
        if any(upload["user"] == user for upload in transfer_uploads):
            supporting_sources.add("file_transfer_audit")
        risk_counts = item["risk_counts"]
        representative_assets = sorted(item["assets"])[:8]
        representative_commands = item["representative_commands"][:8]
        if any("/tmp/" in command.lower() or "chmod 0777" in command.lower() for command in representative_commands):
            judgement = "这组行为更接近文件投放、权限放开与本地执行链，应直接作为重点高风险样本。"
        elif any("curl" in command.lower() and "bash" in command.lower() for command in representative_commands) or any("--private-key" in command for command in representative_commands):
            judgement = "这组行为包含外部脚本直拉执行或明文敏感参数调用，应按高风险下载执行/凭证暴露模式处理。"
        elif any("hot-upgrade" in command.lower() or "patch.sh" in command.lower() for command in representative_commands):
            judgement = "这组行为更像高权限发布或批量补丁操作，风险来自多节点变更与提权执行，而不是典型恶意载荷。"
        elif any("wget" in command.lower() or "installer" in command.lower() for command in representative_commands):
            judgement = "这组行为更接近内部二进制替换、补丁安装或验证链，应归类为高风险运维/发布会话。"
        else:
            judgement = "更像高风险运维或高危调试链，尚不足以单独确认外部入侵。"
        high_risk_accounts.append(
            {
                "user": user,
                "assets": representative_assets,
                "risk_actions": sorted(item["risk_actions"]),
                "risk_counts": risk_counts,
                "representative_commands": representative_commands,
                "supporting_sources": sorted(supporting_sources),
                "judgement": judgement,
            }
        )

    overall_level = "high_risk_pending_review" if cross_source_correlations else "high_risk_operations"
    evidence_provenance = [
        {
            "export_user": event.get("user"),
            "resource_type": event.get("resource_type"),
            "export_time": event.get("event_time"),
            "related_file": event.get("resource"),
            "judgement": "该样本具备明确的导出来源记录。"
        }
        for event in export_events[:8]
    ]
    payload = {
        "sources": [
            {
                "filename": event.asset_context.get("source_file"),
                "source_type": event.source_type,
                "source_role": (
                    "primary_identity_source" if event is latest_login_event else
                    "primary_behavior_source" if event.source_type == "command_audit" else
                    "payload_transfer_source" if event.source_type == "file_transfer_audit" else
                    "control_plane_audit_source" if event.source_type == "operation_audit" else
                    "supplemental_snapshot_only" if event in supplemental_login_events else
                    "supporting_source"
                ),
                "row_count": event.payload.get("row_count", 0),
            }
            for event in raw_events
            if event.source_type in {"login_auth", "command_audit", "file_transfer_audit", "operation_audit"}
        ],
        "supplemental_sources": [
            {
                "filename": event.asset_context.get("source_file"),
                "source_type": event.source_type,
                "source_role": "supplemental_snapshot_only",
                "row_count": event.payload.get("row_count", 0),
            }
            for event in supplemental_login_events
        ],
        "login_summary": {
            "total_rows": len(login_rows),
            "success_count": login_success,
            "failure_count": login_failures,
            "top_failed_accounts": sorted(
                ({"user": user, "count": count} for user, count in failed_accounts.items()),
                key=lambda item: item["count"],
                reverse=True,
            )[:8],
            "proxy_ips": sorted(proxy_ips),
        },
        "command_summary": {
            "total_rows": len(command_rows),
            "effective_command_count": effective_command_count,
            "high_risk_action_counts": command_findings,
            "high_risk_action_labels": {
                command_pattern_labels[key]: value
                for key, value in command_findings.items()
                if value
            },
        },
        "file_transfer_summary": {
            "total_rows": len(transfer_rows),
            "upload_count": len(transfer_uploads),
            "download_count": len(transfer_downloads),
            "high_risk_upload_count": transfer_findings["binary_drop_to_tmp"],
            "data_export_count": transfer_findings["data_export"],
            "top_uploads": transfer_uploads[:8],
        },
        "operation_summary": {
            "total_rows": len(operation_rows),
            "action_counts": operation_actions,
            "create_count": sum(count for action, count in operation_actions.items() if "创建" in action or "create" in action.lower()),
            "update_count": sum(count for action, count in operation_actions.items() if "更新" in action or "update" in action.lower()),
            "export_actions": operation_findings["export_actions"],
            "host_or_account_creation": operation_findings["host_or_account_creation"],
            "authorization_updates": operation_findings["authorization_updates"],
            "session_creation": operation_findings["session_creation"],
            "operator_counts": operator_counts,
            "resource_type_counts": resource_type_counts,
            "export_events": export_events[:10],
            "top_operations": high_impact_operations[:8],
        },
        "cross_source_correlations": cross_source_correlations[:10],
        "high_risk_accounts": high_risk_accounts,
        "evidence_provenance": evidence_provenance,
        "risk_classification": {
            "overall_level": overall_level,
            "category": "jumpserver_multi_source_audit",
            "confidence": "medium" if cross_source_correlations else "medium-low",
            "why_not_direct_intrusion": "当前证据更接近高风险运维与调试链，还不足以单独确认外部入侵已成立。",
            "why_still_high_risk": "已经出现上传、提权、权限放开、执行、服务变更或网络验证等多步高风险操作。",
        },
        "recommended_followups": [
            "按用户、资产、会话和时间窗补齐原始会话级证据，确认是否存在授权维护窗口。",
            "优先复核 /tmp 二进制上传、权限放开、执行与连通性测试链条。",
            "把管理平面里的导出、授权、主机/账号创建动作和主机侧高危命令分开定性，避免用一边替代另一边。",
            "不要直接信任 JumpServer 原始风险等级或代理地址，需结合跨源证据再定性。",
        ],
        "boundary_rules": [
            "do_not_treat_jumpserver_proxy_ip_as_true_attack_source",
            "do_not_trust_native_risk_accept_0",
            "do_not_judge_intrusion_by_single_sudo",
            "tmp_upload_plus_execution_is_high_risk_but_not_auto_malicious",
            "session_level_correlation_is_required",
            "control_plane_actions_support_but_do_not_replace_host_side_execution_evidence",
        ],
    }
    source_files = [event.asset_context.get("source_file") for event in raw_events if event.asset_context.get("source_file")]
    return RawEvent(
        source_type="jumpserver",
        event_hint="jumpserver_multi_source_audit",
        asset_context={
            "asset_id": "jumpserver-multi-source-batch",
            "asset_name": "JumpServer Multi-Source Audit Batch",
            "environment": "unknown",
            "criticality": 4,
            "source_files": source_files,
            "source_type": "jumpserver",
        },
        payload=payload,
    )
